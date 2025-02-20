from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from file_parsers import parse_port_info
from utilities import SEVERITY_ORDER, colored_text

# Define color fills for each severity level
SEVERITY_FILL_COLORS = {
    "Critical": PatternFill(start_color="D7BDE2", end_color="D7BDE2", fill_type="solid"),
    "High": PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid"),
    "Medium": PatternFill(start_color="FAD7A0", end_color="FAD7A0", fill_type="solid"),
    "Low": PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid"),
    "Info": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
}

def sanitize_sheet_name(name):
    """Ensure the sheet name is valid for Excel."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]  


def write_to_excel(data_dict, sheet_name, workbook):
    ws = workbook.create_sheet(title=sheet_name)
    headers = ['Vulnerability Name', 'Severity', 'Affected Hosts', 'Affected Hostnames', 'Recommendations']
    
    # Write headers
    ws.append(headers)
    
    # Sort and write data
    sorted_data = sorted(data_dict.items(), key=lambda x: SEVERITY_ORDER.get(x[1]['Severity'], -1), reverse=True)
    for vuln, data in sorted_data:
        affected_hostnames = []
        
        # Lookup DNS names for each affected host
        for host in data['Affected Hosts']:
            ip_address = host.split(':')[0]  # Extract the IP part of "IP:Port"
            dns_name = next((info['DNS Name'] for info in data.get('Host Info', []) if info['IP Address'] == ip_address), "")
            if dns_name:
                affected_hostnames.append(dns_name)
        
        # Write row to worksheet
        row = [
            vuln,
            data['Severity'],
            ', '.join(data['Affected Hosts']),
            ', '.join(affected_hostnames),
            data['Recommendations'],
        ]
        ws.append(row)
    pass

def write_port_table(nessus_file, workbook, sheet_name, is_external=False):
    
    data = parse_port_info(nessus_file)
    ws = workbook.create_sheet(title=sheet_name)

    if is_external:
        # External columns: DNS, IP, Open Ports, Port Info, URLs
        headers = ["DNS Name", "IP Address", "Open Ports", "Port Info", "URLs"]
        ws.append(headers)
        for host in data:
            ws.append([
                host["DNS Name"],
                host["IP Address"],
                host["Open Ports"],
                host["Port Info"],
                host["URLs"]
            ])
    else:
        # Internal columns: DNS, IP, Open Ports, URLs (no Port Info)
        headers = ["DNS Name", "IP Address", "Open Ports", "URLs"]
        ws.append(headers)
        for host in data:
            ws.append([
                host["DNS Name"],
                host["IP Address"],
                host["Open Ports"],
                host["URLs"]
            ])

    # Auto-adjust column width and enable text wrapping
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (A, B, C, etc.)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)
    pass

def apply_workbook_styling(workbook):
    print(colored_text("Applying consistent styling across all worksheets...", "yellow"))
    header_font = Font(name="Calibri", size=12, bold=True)
    cell_font = Font(name="Calibri", size=11)
    osint_tab_color = "92D050"
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"{colored_text('Styling worksheet:', 'white')} {colored_text(sheet_name, 'green')}")

        # Identify and color OSINT tabs
        if "OSINT" in sheet_name:  # Adjust this condition to match how OSINT tabs are named
            ws.sheet_properties.tabColor = osint_tab_color
            print(f"{colored_text('OSINT tab detected and colored light green:', 'yellow')} {sheet_name}")

        # Wrap text ONLY for External Port Table & Internal Port Table
        # All other sheets: wrap_text = False
        if sheet_name in ["External Port Table", "Internal Port Table"]:
            wrap_text_setting = True
        else:
            wrap_text_setting = False

        # Style headers (first row)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="bottom", 
                wrap_text=wrap_text_setting
            )
            cell.fill = PatternFill(
                start_color="E0E0E0", 
                end_color="E0E0E0", 
                fill_type="solid"
            )
        
        # Style all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = Alignment(
                    horizontal="left", 
                    vertical="bottom", 
                    wrap_text=wrap_text_setting
                )
                
                # Apply severity-based color if in these sheets
                if sheet_name in ["External Scan", "Internal Scan", "Missing Critical Patches"]:
                    # Severity is column B (index = 2)
                    severity = ws.cell(row=cell.row, column=2).value
                    if severity in SEVERITY_FILL_COLORS:
                        cell.fill = SEVERITY_FILL_COLORS[severity]
        
        # Approximate auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        text_length = len(str(cell.value))
                        if text_length > max_length:
                            max_length = text_length
                except Exception as e:
                    print(colored_text(f"Warning when processing {sheet_name}: {e}", "red"))
 
            ws.column_dimensions[column_letter].width = min(max_length + 2, 70)
        
        # Freeze columns A and B if it's External or Internal Scan
        if sheet_name in ["External Scan", "Internal Scan"]:
            ws.freeze_panes = "C1"
    pass
        
        