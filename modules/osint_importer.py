from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ✅ Define headers for each known OSINT sheet
OSINT_HEADERS = {
    "IP Addreses (Whois)": ["IP Ranges", "Registered Name"],
    "Breached Creds (Flare)": ["Username", "Password", "Domain"],
    "Subdomains DNS": ["Subdomain", "IP Address"],
    "Email Addresses": ["Email Address"],
    "Employee Lists": ["Employee", "Description"],
    "Accessible Files": ["URL's", "File Type"],
    "Domains": ["Domains"],
    # Add more as needed...
}

def import_osint_report(osint_file, workbook):
    try:
        osint_workbook = load_workbook(osint_file)

        for sheet_name in osint_workbook.sheetnames:
            osint_sheet = osint_workbook[sheet_name]

            # Create a new sheet name ensuring no duplicates
            base_name = f"OSINT - {sheet_name}"  
            new_sheet_name = base_name
            suffix = 1
            while new_sheet_name in workbook.sheetnames:
                new_sheet_name = f"{base_name}_{suffix}"  
                suffix += 1

            # Create a new sheet in the target workbook
            target_sheet = workbook.create_sheet(title=new_sheet_name)

            # Convert iterator to list so we can inspect data
            rows = list(osint_sheet.iter_rows(values_only=True))  

            if not rows:  # If the OSINT sheet is empty, skip it
                print(f"Warning: OSINT sheet '{sheet_name}' is empty. Skipping.")
                continue 

            # ✅ Ensure headers are always written
            if sheet_name in OSINT_HEADERS:
                headers = OSINT_HEADERS[sheet_name]  # Use predefined headers
            else:
                num_columns = max(len(row) for row in rows) if rows else 0
                headers = [f"Column {chr(65 + i)}" for i in range(num_columns)]  # Generate "Column A, B, C..."

            # ✅ Always write headers first
            target_sheet.append(headers)
            for col, header in enumerate(headers, start=1):
                cell = target_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

            # ✅ Copy all remaining data rows (do NOT assume first row is a header)
            for row in rows:  
                target_sheet.append(row)

            # ✅ Auto-adjust column widths
            for column in target_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  
                target_sheet.column_dimensions[column_letter].width = adjusted_width

            print(f"Imported sheet: {sheet_name} as {new_sheet_name} with {target_sheet.max_row} rows.")

        print(f"Successfully imported OSINT report: {osint_file}")

    except Exception as e:
        print(f"Error importing OSINT report: {e}")
