#!/usr/bin/env python3
import xml.etree.ElementTree as ET
import argparse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
import time
import threading
from itertools import cycle
import re
from threading import Event
import json
import zipfile

try:
    import ipaddress  # For detecting IPv6 if needed
except ImportError:
    ipaddress = None

# Mapping for severity levels
SEVERITY_MAP = {
    "0": "Informational",
    "1": "Low",
    "2": "Medium",
    "3": "High",
    "4": "Critical"
}

SEVERITY_ORDER = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1, "Informational": 0}

# Define color fills for each severity level
SEVERITY_FILL_COLORS = {
    "Critical": PatternFill(start_color="D7BDE2", end_color="D7BDE2", fill_type="solid"),
    "High": PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid"),
    "Medium": PatternFill(start_color="FAD7A0", end_color="FAD7A0", fill_type="solid"),
    "Low": PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid"),
    "Info": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
}

def sort_worksheet_by_severity(ws, severity_column="B"):
    # Convert column letter to index
    severity_col_index = ws[severity_column][0].column

    # Read all rows into a list, skip the header
    data = list(ws.iter_rows(values_only=True))
    header, rows = data[0], data[1:]

    # Sort rows by the severity column (assuming severity is in text format, e.g., 'Critical', 'High', etc.)
    severity_order = {"Critical": 1, "High": 2, "Medium": 3, "Low": 4, "Info": 5}
    rows.sort(key=lambda row: severity_order.get(row[severity_col_index - 1], 6))

    # Clear the worksheet and write the sorted rows back
    for row_index, row in enumerate([header] + rows, start=1):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

def colored_text(text, color="white"):
    colors = {
        "red": "\033[91m",
        "green": "\033[92m",
        "yellow": "\033[93m",
        "blue": "\033[94m",
        "white": "\033[97m",
        "reset": "\033[0m",
    }
    return f"{colors.get(color, colors['white'])}{text}{colors['reset']}"


class SpinnerManager:
    def __init__(self, msg):
        self.stop_event = Event()
        self.msg = colored_text(msg, "blue")  # Make the message blue
        self.spinner_thread = None

    def spin(self):
        spinner_cycle = cycle(['|', '/', '-', '\\'])
        while not self.stop_event.is_set():
            print(f"\r{self.msg} {colored_text(next(spinner_cycle), 'blue')}   ", end="", flush=True)  # Extra spaces after the spinner
            time.sleep(0.1)
        print("\r" + colored_text(f"{self.msg}   Completed!", "green"))  # Extra spaces before 'Completed!'

    def __enter__(self):
        self.spinner_thread = threading.Thread(target=self.spin)
        self.spinner_thread.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop_event.set()
        if self.spinner_thread:
            self.spinner_thread.join()



def parse_nessus_file(file_path):
    """Parse .nessus file and extract vulnerability data, skipping 'Informational' findings."""
    try:
        print(f"Parsing Nessus file: {file_path}")
        print("\r" + colored_text(f"Parsing Nessus File: {file_path}", "green"), end="")
        tree = ET.parse(file_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"Error: Unable to parse the .nessus file: {file_path}.\nDetails: {e}")
        return {}

    vuln_dict = {}
    report_hosts = list(root.iter('ReportHost'))
    print(f"Found {len(report_hosts)} hosts in the Nessus file.")

    for host in report_hosts:
        hostname = host.attrib.get('name', '')
        for block in host.iter('ReportItem'):
            plugin_name = block.attrib.get('pluginName', '')
            severity = block.attrib.get('severity', '0')
            severity_label = SEVERITY_MAP.get(severity, "Unknown")

            # Skip all 'Informational' findings:
            if severity_label == "Informational":
                continue

            host_port = f"{hostname}:{block.attrib.get('port', '')}"

            if plugin_name not in vuln_dict:
                vuln_dict[plugin_name] = {
                    'Severity': severity_label,
                    'Affected Hosts': [host_port],
                    'Recommendations': block.findtext('solution', ''),
                    'Description': block.findtext('description', '')
                }
            else:
                if host_port not in vuln_dict[plugin_name]['Affected Hosts']:
                    vuln_dict[plugin_name]['Affected Hosts'].append(host_port)

    print(f"Parsed {len(vuln_dict)} vulnerabilities from the Nessus file.")
    return vuln_dict

def write_to_excel(data_dict, sheet_name, workbook):
    """Write parsed data to Excel worksheet (External Scan or Internal Scan)."""
    print(f"Writing data to the '{sheet_name}' sheet...")
    ws = workbook.create_sheet(title=sheet_name)
    headers = ['Vulnerability Name', 'Severity', 'Affected Hosts', 'Recommendations']
    
    # Write headers
    ws.append(headers)
    
    # Sort and write data
    sorted_data = sorted(data_dict.items(), key=lambda x: SEVERITY_ORDER.get(x[1]['Severity'], -1), reverse=True)
    for vuln, data in sorted_data:
        row = [
            vuln,
            data['Severity'],
            ', '.join(data['Affected Hosts']),
            data['Recommendations'],
        ]
        ws.append(row)

def write_processes_tab(file_path, workbook):
    """Write remote access tool processes information to Excel worksheet."""
    print("Processing remote access processes...")
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        ws = workbook.create_sheet(title="Remote Access Processes")
        headers = ['IP Address', 'Hostname', 'Remote Access Processes']
        ws.append(headers)
        
        # Define target processes to look for
        target_processes = [
            "TeamViewer", "remote", "Mikogo", "screenconnect", "atera", "goto", 
            "ISL Online", "LogMeIn", "LogMeIn Pro", "LogMeIn Central", 
            "LogMeInRescue", "GoToMyPC", "Dameware Remote Support", 
            "Dameware Remote Everywhere", "Dameware Mini Remote Control", 
            "Parallels", "Mikogo"
        ]

        results = {}

        for ReportHost in root.findall('.//ReportHost'):
            # Get IP address
            ip_address = ReportHost.find('HostProperties/tag[@name="host-ip"]')
            ip_address = ip_address.text if ip_address is not None else "Unknown IP"
            
            hostname = "Unknown Hostname"
            for host_property in ReportHost.findall('HostProperties/tag'):
                if host_property.get('name') == 'host-fqdn':
                    hostname = host_property.text
                    break

            # Look for process information
            for ReportItem in ReportHost.findall('.//ReportItem'):
                plugin_name = ReportItem.get('pluginName')
                if plugin_name == "Microsoft Windows Process Information":
                    process_info = ReportItem.find('plugin_output')
                    if process_info is not None:
                        found_processes = re.findall(r'[\w.-]+\.exe', process_info.text, re.IGNORECASE)
                        for found_process in found_processes:
                            for process in target_processes:
                                if process.lower() in found_process.lower():
                                    if ip_address not in results:
                                        results[ip_address] = {'hostname': hostname, 'processes': set()}
                                    results[ip_address]['processes'].add(found_process)

        # Write results to Excel
        for ip, info in sorted(results.items()):
            process_list = ', '.join(info['processes'])
            ws.append([ip, info['hostname'], process_list])

    except Exception as e:
        print(f"Error processing processes tab: {e}")

def write_protocols_tab(file_path, workbook):
    """
    Write protocol information to Excel worksheet, focusing on:
      1) LLMNR (port 5355)
      2) mDNS  (port 5353)
      3) NBT-NS (port 137)
      4) IPv6  (detected via IP address format)
    """
    print("Processing protocols from Nessus file...")
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        ws = workbook.create_sheet(title="Protocols")
        headers = ['Protocol', 'Host Count', 'Hosts']
        ws.append(headers)
        
        # We'll detect the protocols by port:
        port_map = {
            "5355": "LLMNR",
            "5353": "mDNS",
            "137":  "NBT-NS",
        }
        
        protocol_dict = {
            "LLMNR": set(),
            "mDNS": set(),
            "NBT-NS": set(),
            "IPv6": set(),
        }
        
        for ReportHost in root.findall('.//ReportHost'):
            # Grab IP address
            ip_address = ReportHost.find('HostProperties/tag[@name="host-ip"]')
            ip_address = (
                ip_address.text if ip_address is not None
                else ReportHost.attrib.get('name', 'Unknown IP')
            )
            
            # Check if the IP is IPv6:
            if ipaddress:
                try:
                    ip_obj = ipaddress.ip_address(ip_address)
                    if ip_obj.version == 6:
                        protocol_dict["IPv6"].add(ip_address)
                except ValueError:
                    pass

            # For each ReportItem, check if the port is in port_map
            for ReportItem in ReportHost.findall('.//ReportItem'):
                port = ReportItem.get('port', '')
                if port in port_map:
                    protocol_name = port_map[port]
                    protocol_dict[protocol_name].add(ip_address)
        
        # Now write them out in the order: LLMNR, mDNS, NBT-NS, IPv6
        for protocol_name in ["LLMNR", "mDNS", "NBT-NS", "IPv6"]:
            hosts = sorted(protocol_dict[protocol_name])
            if hosts:  # Only add rows for protocols that actually exist
                ws.append([
                    protocol_name,
                    len(hosts),
                    ', '.join(hosts)
                ])
        
        print("Finished processing protocols.")
        
    except Exception as e:
        print(f"Error processing protocols tab: {e}")

def write_unsupported_software(data_dict, workbook):
    """
    Write unsupported software information to Excel worksheet with columns:
      Software Name, Severity, Host Count, Affected Hosts
    Sorted by host count descending.
    """
    ws = workbook.create_sheet(title="Unsupported Software")
    
    headers = ['Software Name', 'Severity', 'Host Count', 'Affected Hosts']
    ws.append(headers)
    
    # Filter for "unsupported" or "end of life"
    unsupported_software = {
        name: data for name, data in data_dict.items()
        if 'unsupported' in name.lower() or 'end of life' in name.lower()
    }
    
    # Sort by the length of the affected hosts list (descending)
    sorted_unsupported_software = sorted(
        unsupported_software.items(),
        key=lambda x: len(x[1]['Affected Hosts']),
        reverse=True
    )
    
    for name, data in sorted_unsupported_software:
        host_count = len(data['Affected Hosts'])
        ws.append([
            name,
            data['Severity'],
            host_count,
            ', '.join(data['Affected Hosts'])
        ])

def write_missing_critical_patches(data_dict, workbook):
    """
    Write 'Missing Critical Patches' information to a new worksheet with columns:
      1) Patch         (A)
      2) Severity      (B)
      3) Host Count    (C)
      4) Affected Hosts(D)

    This includes only severity = [Medium, High, Critical].
    Sorts primarily by severity (Critical > High > Medium),
    then secondarily by host count (descending).
    """
    ws = workbook.create_sheet(title="Missing Critical Patches")
    
    # Create headers
    headers = ['Patch', 'Severity', 'Host Count', 'Affected Hosts']
    ws.append(headers)
    
    # Filter for medium, high, and critical
    allowed_severities = {"Medium", "High", "Critical"}
    missing_patches = {
        name: data for name, data in data_dict.items()
        if data["Severity"] in allowed_severities
    }
    
    # Sort by severity first, then by host count
    sorted_patches = sorted(
        missing_patches.items(),
        key=lambda x: (
            SEVERITY_ORDER.get(x[1]["Severity"], 0),  # severity as primary
            len(x[1]['Affected Hosts'])               # host count as secondary
        ),
        reverse=True
    )
    
    for name, data in sorted_patches:
        host_count = len(data['Affected Hosts'])
        ws.append([
            name,              # Patch
            data['Severity'], # Severity
            host_count,        # Host Count
            ', '.join(data['Affected Hosts'])
        ])

def is_valid_customization(file_path):
    """Check if the provided file path is a valid .customization file."""
    return os.path.isfile(file_path) and file_path.endswith('.customization')

def append_customization_data(custom_file, workbook):
    """
    Loads a .customization (JSON) file and appends the vulnerabilities
    into the existing 'External Scan' and 'Internal Scan' sheets.

    Columns appended:
      1) Vulnerability Name
      2) Severity
      3) Affected Hosts
      4) Recommendations

    If either sheet doesn't exist, we'll simply warn and not append.
    """
    print(f"\nLoading customization from {custom_file}")
    with open(custom_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    vulns = data.get("vulnerabilities", [])
    external_rows = []
    internal_rows = []

    # Separate external vs internal
    for vuln in vulns:
        title = vuln.get("title", "Unknown Title")
        severity = vuln.get("priority", "Unknown Severity")
        recommendation = vuln.get("remediation_recommendation", "")
        
        # Build the Affected Hosts string
        affected_assets = vuln.get("affected_assets", [])
        hosts = []
        is_internal = False

        for asset_entry in affected_assets:
            host_name = asset_entry.get("asset", "Unknown Host")
            hosts.append(host_name)
            # Check tags to see if it's internal
            for tag_dict in asset_entry.get("assetCustomTags", []):
                if tag_dict.get("Internal") == "true":
                    is_internal = True

        row_data = [
            title,                   # Vulnerability Name
            severity,               # Severity
            ", ".join(hosts),       # Affected Hosts
            recommendation          # Recommendations
        ]
        
        if is_internal:
            internal_rows.append(row_data)
        else:
            external_rows.append(row_data)

    # Append to "External Scan"
    if "External Scan" in workbook.sheetnames:
        ws_ext = workbook["External Scan"]
        for row in external_rows:
            ws_ext.append(row)
        print(f"{colored_text('Appended', 'white')} {colored_text(len(external_rows), 'green')} {colored_text('vulnerabilities', 'white')} to {colored_text('External Scan', 'green')} tab.")
    else:
        print(colored_text("Warning: 'External Scan' sheet not found. No external data appended.", "red"))

    # Append to "Internal Scan"
    if "Internal Scan" in workbook.sheetnames:
        ws_int = workbook["Internal Scan"]
        for row in internal_rows:
            ws_int.append(row)
        print(f"{colored_text('Appended', 'white')} {colored_text(len(internal_rows), 'green')} {colored_text('vulnerabilities', 'white')} to {colored_text('Internal Scan', 'green')} tab.")
    else:
        print(colored_text("Warning: 'Internal Scan' sheet not found. No Internal data appended.", "red"))

def is_valid_zip(file_path):
    """Check if the provided file path is a valid .zip file."""
    return os.path.isfile(file_path) and file_path.endswith('.zip')
        
def parse_bloodhound_zip(zip_path, workbook):
    if not is_valid_zip(zip_path):
        print(f"Error: {zip_path} is not a valid .zip file.")
        return

    print(f"Processing BloodHound file: {zip_path}")

    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_file:
            # Look for `users.json` or similar files in the archive
            for file_name in zip_file.namelist():
                if "users.json" in file_name.lower():
                    print(f"Found: {file_name}")
                    with zip_file.open(file_name) as f:
                        data = json.load(f)

                        # Define detection rules
                        detection_rules = {
                            "Kerberoastable Users": {
                                "filter": lambda user: user.get('Properties', {}).get('hasspn', False),
                                "headers": ['Username', 'Display Name', 'SPN', 'Description'],
                                "extract": lambda user: [
                                    user.get('Properties', {}).get('samaccountname', 'N/A'),
                                    user.get('Properties', {}).get('displayname', 'N/A'),
                                    ', '.join(user.get('Properties', {}).get('serviceprincipalnames', [])) or 'N/A',
                                    user.get('Properties', {}).get('description', 'N/A')
                                ]
                            },
                            "ASREPRoastable Users": {
                                "filter": lambda user: user.get('Properties', {}).get('dontreqpreauth', False),
                                "headers": ['Username', 'Display Name', 'Description'],
                                "extract": lambda user: [
                                    user.get('Properties', {}).get('samaccountname', 'N/A'),
                                    user.get('Properties', {}).get('displayname', 'N/A'),
                                    user.get('Properties', {}).get('description', 'N/A')
                                ]
                            },
                            "Users With Unconstrained Delegation": {
                                "filter": lambda user: user.get('Properties', {}).get('unconstraineddelegation', False),
                                "headers": ['Username', 'Display Name', 'Description'],
                                "extract": lambda user: [
                                    user.get('Properties', {}).get('samaccountname', 'N/A'),
                                    user.get('Properties', {}).get('displayname', 'N/A'),
                                    user.get('Properties', {}).get('description', 'N/A')
                                ]
                            },
                            "Users With RBCD": {
                                "filter": lambda user: user.get('Properties', {}).get('rbcdenabled', False),
                                "headers": ['Username', 'Display Name', 'Description'],
                                "extract": lambda user: [
                                    user.get('Properties', {}).get('samaccountname', 'N/A'),
                                    user.get('Properties', {}).get('displayname', 'N/A'),
                                    user.get('Properties', {}).get('description', 'N/A')
                                ]
                            }
                        }

                        # Process each detection rule
                        for sheet_name, rule in detection_rules.items():
                            rows = []
                            for user in data.get('data', []):
                                if rule["filter"](user):
                                    rows.append(rule["extract"](user))

                            if rows:
                                print(f"Creating sheet: {sheet_name}")
                                ws = workbook.create_sheet(title=sheet_name)
                                ws.append(rule["headers"])
                                for row in rows:
                                    ws.append(row)

        print("Finished processing BloodHound file.")
    except Exception as e:
        print(f"Error processing BloodHound file: {e}")


def sanitize_sheet_name(name):
    """Ensure the sheet name is valid for Excel."""
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]  # Limit to 31 characters

def apply_workbook_styling(workbook):
    """Apply consistent styling across all worksheets in the workbook."""
    print(colored_text("Applying consistent styling across all worksheets...", "yellow"))
    header_font = Font(name="Calibri", size=12, bold=True)
    cell_font = Font(name="Calibri", size=11)
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"{colored_text('Styling worksheet:', 'white')} {colored_text(sheet_name, 'green')}")


        # Wrap text ONLY for External Port Table & Internal Port Table
        # All other sheets: wrap_text = False
        if sheet_name in ["External Port Table", "Internal Port Table"]:
            wrap_text_setting = True
        else:
            wrap_text_setting = False

        # Style headers (first row)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="left", 
                vertical="bottom", 
                wrap_text=wrap_text_setting
            )
            cell.fill = PatternFill(
                start_color="E0E0E0", 
                end_color="E0E0E0", 
                fill_type="solid"
            )
        
        # Style all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = Alignment(
                    horizontal="left", 
                    vertical="bottom", 
                    wrap_text=wrap_text_setting
                )
                
                # Apply severity-based color if in these sheets
                if sheet_name in ["External Scan", "Internal Scan", "Missing Critical Patches"]:
                    # Severity is column B (index = 2)
                    severity = ws.cell(row=cell.row, column=2).value
                    if severity in SEVERITY_FILL_COLORS:
                        cell.fill = SEVERITY_FILL_COLORS[severity]
        
        # Approximate auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        text_length = len(str(cell.value))
                        if text_length > max_length:
                            max_length = text_length
                except Exception as e:
                    print(f"Warning when processing {sheet_name}: {e}")
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 70)
        
        # Freeze columns A and B if it's External or Internal Scan
        if sheet_name in ["External Scan", "Internal Scan"]:
            ws.freeze_panes = "C1"

def parse_port_info(nessus_file):
    """
    Parse a Nessus XML file and return a list of dictionaries with:
      DNS Name, IP Address, Open Ports, Port Info, and a computed 'URLs' field.
    """
    tree = ET.parse(nessus_file)
    root = tree.getroot()

    host_info = []

    # Iterate through each ReportHost
    for report_host in root.iter("ReportHost"):
        host_ip = ""
        dns_name = ""
        tcp_ports = set()
        udp_ports = set()
        port_details = {}

        # Extract the host IP and DNS name
        for tag in report_host.iter("tag"):
            if tag.attrib.get("name") == "host-ip":
                host_ip = tag.text
            elif tag.attrib.get("name") == "host-fqdn":
                dns_name = tag.text

        # Iterate through ReportItems to find open ports
        for item in report_host.iter("ReportItem"):
            port = item.attrib["port"]
            protocol = item.attrib["protocol"]
            svc_name = item.attrib["svc_name"]

            if port != "0":
                if protocol.lower() == "tcp":
                    tcp_ports.add(port)
                elif protocol.lower() == "udp":
                    udp_ports.add(port)

                # Example re-label for ike
                if svc_name == "ike":
                    svc_name = "IPSEC Ike"

                if "?" in svc_name:
                    port_details[port] = f"Unable to fingerprint a running service on Port {port}."
                else:
                    if svc_name == "www":
                        port_details[port] = (
                            f"Port {port} appears to be associated with a web server..."
                        )
                    else:
                        port_details[port] = f"Port {port} appears to be associated with {svc_name}."

        # Build the "Open Ports" string
        parts = []
        if tcp_ports:
            parts.append("TCP: " + ", ".join(sorted(tcp_ports)))
        if udp_ports:
            parts.append("UDP: " + ", ".join(sorted(udp_ports)))
        open_ports_str = "\n".join(parts)

        # Combine per-port messages (Port Info)
        port_info_str = "\n".join(port_details.values())

        # Build URLs
        urls = set()
        # If we see port 80 in TCP, assume http
        if "80" in tcp_ports:
            if dns_name:
                urls.add(f"http://{dns_name}")
            else:
                urls.add(f"http://{host_ip}")
        # If we see port 443 in TCP, assume https
        if "443" in tcp_ports:
            if dns_name:
                urls.add(f"https://{dns_name}")
            else:
                urls.add(f"https://{host_ip}")

        url_str = "\n".join(sorted(urls))

        host_info.append({
            "DNS Name": dns_name or "",
            "IP Address": host_ip or "",
            "Open Ports": open_ports_str,
            "Port Info": port_info_str,
            "URLs": url_str
        })
    
    return host_info

def write_smb_signing_off(file_path, workbook):
    """
    Parse the internal Nessus file for hosts that have SMB signing turned off or not required,
    then create a new tab "SMB Signing Off" with columns:
      [Total Count | DNS Name | IP Address]

    We only show the total count once, in the top-left cell (row 2, col A).
    """

    print("Processing SMB signing off hosts...")

    # We'll look for pluginName containing "SMB Signing" + "Not Required"
    SMB_MATCH_STRING = "smb signing"
    NOT_REQ_STRING = "not required"

    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
    except Exception as e:
        print(f"Error parsing Nessus file in write_smb_signing_off: {e}")
        return

    results = []  # Store tuples of (dns_name, ip_address)

    for report_host in root.iter("ReportHost"):
        ip_address = "Unknown IP"
        dns_name = "Unknown DNS"
        
        # Extract IP & DNS if available
        for tag in report_host.iter("tag"):
            if tag.attrib.get("name") == "host-ip":
                ip_address = tag.text
            elif tag.attrib.get("name") == "host-fqdn":
                dns_name = tag.text

        # Check each ReportItem for a matching plugin
        for item in report_host.iter("ReportItem"):
            plugin_name = item.attrib.get("pluginName", "")
            if (
                SMB_MATCH_STRING in plugin_name.lower()
                and NOT_REQ_STRING in plugin_name.lower()
            ):
                results.append((dns_name, ip_address))
                # Break so we don't list the same host multiple times
                break

    # Create a new sheet
    ws = workbook.create_sheet(title="SMB Signing Off")

    # Headers
    headers = ["Total Count", "DNS Name", "IP Address"]
    ws.append(headers)

    total_count = len(results)

    if total_count > 0:
        # Put the total count in the first column (row 2), leave DNS/IP blank
        ws.append([total_count, "", ""])
        # For each host, put DNS (col B) and IP (col C), and leave col A empty
        for (dns, ip) in results:
            ws.append(["", dns, ip])
    else:
        # If no hosts found, just show "0" in the second row
        ws.append([0, "", ""])

    print(f"Found {total_count} host(s) with SMB signing off.")

def write_port_table(nessus_file, workbook, sheet_name, is_external=False):
    """
    Creates a new sheet in 'workbook' named sheet_name and populates it with:
    - For external scans:
       DNS Name | IP Address | Open Ports | Port Info | URLs
    - For internal scans:
       DNS Name | IP Address | Open Ports | URLs  (No Port Info)
    """
    data = parse_port_info(nessus_file)
    ws = workbook.create_sheet(title=sheet_name)

    if is_external:
        # External columns: DNS, IP, Open Ports, Port Info, URLs
        headers = ["DNS Name", "IP Address", "Open Ports", "Port Info", "URLs"]
        ws.append(headers)
        for host in data:
            row = [
                host["DNS Name"],
                host["IP Address"],
                host["Open Ports"],
                host["Port Info"],
                host["URLs"]
            ]
            ws.append(row)
    else:
        # Internal columns: DNS, IP, Open Ports, URLs (no Port Info)
        headers = ["DNS Name", "IP Address", "Open Ports", "URLs"]
        ws.append(headers)
        for host in data:
            row = [
                host["DNS Name"],
                host["IP Address"],
                host["Open Ports"],
                host["URLs"]
            ]
            ws.append(row)

def main():
    parser = argparse.ArgumentParser(description='Parse Nessus files and generate Excel report.')
    parser.add_argument('-e', '--external', help='Path to the external Nessus scan file.')
    parser.add_argument('-i', '--internal', help='Path to the internal Nessus scan file.')
    parser.add_argument('-a', '--attackforge', help='Path to the .customization (AttackForge) file.')
    parser.add_argument('-b', '--bloodhound', help='Path to the BloodHound .zip file.')
    parser.add_argument('-o', '--output', default='Nessus_Report.xlsx',
                        help='Output Excel file name (default: Nessus_Report.xlsx)')

    args = parser.parse_args()

    if not any([args.external, args.internal, args.attackforge, args.bloodhound]):
        parser.error("Please provide at least one input file.")

    # Create a new workbook
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # Process the provided files
    process_files(args, workbook)

def process_files(args, workbook):
    with SpinnerManager(colored_text("Processing files", "blue")):
        # EXTERNAL
        if args.external:
            if not os.path.isfile(args.external) or not args.external.lower().endswith('.nessus'):
                print(colored_text("Error: External file is not a valid .nessus file.", "red"))
            else:
                external_data = parse_nessus_file(args.external)
                write_to_excel(external_data, 'External Scan', workbook)
                write_port_table(args.external, workbook, "External Port Table", is_external=True)

        # INTERNAL
        if args.internal:
            if not os.path.isfile(args.internal) or not args.internal.lower().endswith('.nessus'):
                print(colored_text("Error: Internal file is not a valid .nessus file.", "red"))
            else:
                internal_data = parse_nessus_file(args.internal)
                write_to_excel(internal_data, 'Internal Scan', workbook)
                write_processes_tab(args.internal, workbook)
                write_unsupported_software(internal_data, workbook)
                write_missing_critical_patches(internal_data, workbook)
                write_protocols_tab(args.internal, workbook)
                write_port_table(args.internal, workbook, "Internal Port Table", is_external=False)
                write_smb_signing_off(args.internal, workbook)

        # BLOODHOUND
        if args.bloodhound:
            if is_valid_zip(args.bloodhound):
                parse_bloodhound_zip(args.bloodhound, workbook)
            else:
                print(colored_text("Error: Provided BloodHound file is not a valid .zip file.", "red"))

        # ATTACKFORGE
        if args.attackforge:
            if is_valid_customization(args.attackforge):
                append_customization_data(args.attackforge, workbook)
            else:
                print(colored_text("Error: Provided AttackForge file is not a valid .customization file.", "red"))
                
        # Sort External Scan and Internal Scan sheets by severity
        if "External Scan" in workbook.sheetnames:
            sort_worksheet_by_severity(workbook["External Scan"])

        if "Internal Scan" in workbook.sheetnames:
            sort_worksheet_by_severity(workbook["Internal Scan"])

        # Apply styling to all sheets after they're created
        apply_workbook_styling(workbook)

    try:
        workbook.save(args.output)
        print(f"\n{colored_text('Report saved successfully!!', 'white')}: {colored_text(args.output, 'green')}")
    except Exception as e:
        print(f"\nError saving report: {e}")



if __name__ == "__main__":
    main()
